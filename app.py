import customtkinter as ctk
from tkinter import filedialog, messagebox
from PIL import Image as PILImage
import openpyxl
from openpyxl.styles import PatternFill, Font
import urllib.request
import threading
import os
import sys
import ssl
import certifi
import zipfile
import shutil
import tempfile
import time
import requests
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Image as RLImage, Paragraph
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
    DND_AVAILABLE = True
except Exception:
    DND_AVAILABLE = False

# ── Version ───────────────────────────────────────────────
CURRENT_VERSION = "1.0.9"
VERSION_URL = "https://raw.githubusercontent.com/MrSamuelB/schedule-report-payroll/refs/heads/main/version.txt"
DOWNLOAD_URL_WIN = "https://github.com/MrSamuelB/schedule-report-payroll/releases/latest/download/ScheduleReportForPayroll-Windows.zip"
DOWNLOAD_URL_MAC = "https://github.com/MrSamuelB/schedule-report-payroll/releases/latest/download/ScheduleReportForPayroll-Mac.zip"

def get_download_url():
    if sys.platform == "win32":
        return DOWNLOAD_URL_WIN
    return DOWNLOAD_URL_MAC

def get_ssl_context():
    try:
        ctx = ssl.create_default_context(cafile=certifi.where())
        return ctx
    except Exception:
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        return ctx

# ── Rules ─────────────────────────────────────────────────
COLUMNS_TO_DELETE = [
    "medical record number", "branch name", "phone", "task type",
    "address", "city", "state", "zip code", "employee id",
    "time in", "time out", "documentation time", "documentation time (min)",
    "travel time", "travel time (min)", "insurance", "productivity units"
]

TASK_NAMES_TO_DELETE = [
    "30-day summary", "additional hospital records", "admit summary",
    "case conference and 60 day summary", "cms 485", "cms 486", "consent", "cota sup",
    "pta sup", "lpn sup", "discharge summary", "dnr", "dpa",
    "dpoa", "emergency plan", "follow up vpoc", "f/up vpoc", "f2f encounter",
    "f2f notes", "frequency order", "hospitalization", "insurance payment",
    "insurance verification", "lab results", "medication profile",
    "msw communication", "msw comm", "nomnc", "otcomm", "ot comm", "ovn",
    "patient communication", "physician order", "ptcomm", "pt comm",
    "referral approval", "referral", "release of information form", "orc",
    "roc", "roc docs", "rocdocs", "soc email", "transfer summary",
    "wound company ovn", "aide care plan"
]

STATUS_HIGHLIGHT = ["not started", "saved"]

RED_FONT = Font(color="FF0000")
RED_FILL = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")

# ── Logo path ─────────────────────────────────────────────
def resource_path(filename):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, filename)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)

LOGO_PATH = resource_path("Logo.png")

# ── Auto-update ───────────────────────────────────────────
def download_and_relaunch():
    try:
        download_url = get_download_url()
        tmp_dir = tempfile.mkdtemp()
        zip_path = os.path.join(tmp_dir, "update.zip")

        app.after(0, lambda: status_label.configure(text="Debug: Starting download..."))

        response = requests.get(download_url, stream=True, verify=certifi.where())
        response.raise_for_status()

        with open(zip_path, "wb") as f:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)

        app.after(0, lambda: status_label.configure(text="Debug: Download complete. Extracting..."))
        time.sleep(1)

        extract_dir = os.path.join(tmp_dir, "extracted")
        os.makedirs(extract_dir, exist_ok=True)
        with zipfile.ZipFile(zip_path, "r") as z:
            z.extractall(extract_dir)

        app.after(0, lambda: status_label.configure(text=f"Debug: Extracted. Contents: {os.listdir(extract_dir)}"))
        time.sleep(2)

        if sys.platform == "win32":
            new_exe = None
            for root, dirs, files in os.walk(extract_dir):
                for f in files:
                    if f.endswith(".exe"):
                        new_exe = os.path.join(root, f)
                        break
            if new_exe:
                current_exe = sys.executable
                batch = os.path.join(tmp_dir, "update.bat")
                with open(batch, "w") as f:
                    f.write(f"""@echo off
timeout /t 2 /nobreak
copy /y "{new_exe}" "{current_exe}"
start "" "{current_exe}"
""")
                import subprocess
                subprocess.Popen(["cmd", "/c", batch], creationflags=0x08000000)
                app.after(0, app.quit)

        else:
            import subprocess
            new_app_path = None
            for item in os.listdir(extract_dir):
                if item.endswith(".app"):
                    new_app_path = os.path.join(extract_dir, item)
                    break

            if not new_app_path:
                app.after(0, lambda: status_label.configure(text="Debug: No .app found in zip"))
                return

            app.after(0, lambda: status_label.configure(text=f"Debug: Found app: {os.path.basename(new_app_path)}"))
            time.sleep(1)

            app_name = os.path.basename(new_app_path)
            downloads = os.path.expanduser("~/Downloads")
            staged = os.path.join(downloads, app_name)

            app.after(0, lambda: status_label.configure(text=f"Debug: Copying to Downloads..."))
            time.sleep(1)

            try:
                if os.path.exists(staged):
                    shutil.rmtree(staged)
                shutil.copytree(new_app_path, staged)
                app.after(0, lambda: status_label.configure(text="Debug: Copy done! Opening new app..."))
                time.sleep(1)
            except Exception as copy_err:
                app.after(0, lambda: status_label.configure(text=f"Debug copy error: {str(copy_err)}"))
                return

            try:
                subprocess.Popen(["open", staged])
                app.after(0, lambda: status_label.configure(text="Debug: Open called! Quitting old app..."))
                time.sleep(1)
                app.after(0, app.quit)
            except Exception as open_err:
                app.after(0, lambda: status_label.configure(text=f"Debug open error: {str(open_err)}"))

    except Exception as e:
        app.after(0, lambda: status_label.configure(text=f"Update failed: {str(e)}"))

def check_for_updates():
    try:
        ctx = get_ssl_context()
        with urllib.request.urlopen(VERSION_URL, timeout=5, context=ctx) as response:
            latest_version = response.read().decode("utf-8").strip()

        if latest_version != CURRENT_VERSION:
            def show_update_popup():
                update_win = ctk.CTkToplevel(app)
                update_win.title("Update Available")
                update_win.geometry("400x220")
                update_win.resizable(False, False)
                update_win.grab_set()

                ctk.CTkLabel(
                    update_win,
                    text="A new version is available!",
                    font=ctk.CTkFont(size=16, weight="bold")
                ).pack(pady=(30, 5))

                ctk.CTkLabel(
                    update_win,
                    text=f"Version {latest_version} is ready.\nClick below to update automatically.",
                    font=ctk.CTkFont(size=13),
                    text_color="gray"
                ).pack(pady=5)

                def do_update():
                    update_win.destroy()
                    status_label.configure(text="Starting update...")
                    app.update()
                    threading.Thread(target=download_and_relaunch, daemon=True).start()

                ctk.CTkButton(
                    update_win,
                    text="Update Now",
                    font=ctk.CTkFont(size=14),
                    height=45,
                    width=220,
                    command=do_update
                ).pack(pady=20)

            app.after(0, show_update_popup)

    except Exception:
        pass

# ── Helpers ───────────────────────────────────────────────
def get_col_index(ws, name):
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value and str(cell.value).strip().lower() == name.lower():
            return col_idx
    return None

def move_column_to(ws, from_idx, to_idx):
    if from_idx == to_idx:
        return
    col_data = [row[from_idx - 1].value for row in ws.iter_rows()]
    ws.delete_cols(from_idx)
    ws.insert_cols(to_idx)
    for row_idx, value in enumerate(col_data, start=1):
        ws.cell(row=row_idx, column=to_idx, value=value)

# ── Process data ──────────────────────────────────────────
def process_data(ws):
    # Step 0: Sort by Target Date chronologically
    target_date_idx = get_col_index(ws, "Target Date")
    if target_date_idx:
        data_rows = [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]
        def parse_date(row):
            val = str(row[target_date_idx - 1] or "").strip()
            try:
                return datetime.strptime(val, "%m/%d/%Y")
            except:
                return datetime.max
        data_rows.sort(key=parse_date)
        for row_idx, row_data in enumerate(data_rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    # Step 1: Delete specified columns
    cols_to_remove = []
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value and str(cell.value).strip().lower() in COLUMNS_TO_DELETE:
            cols_to_remove.append(col_idx)
    for col_idx in sorted(cols_to_remove, reverse=True):
        ws.delete_cols(col_idx)

    # Step 2: Move Provider Last → col B, Provider First → col C
    provider_last_idx = get_col_index(ws, "Provider Last")
    if provider_last_idx:
        move_column_to(ws, provider_last_idx, 2)
    provider_first_idx = get_col_index(ws, "Provider First")
    if provider_first_idx:
        move_column_to(ws, provider_first_idx, 3)

    # Step 3: Sort by Task Name alphabetically
    task_name_idx = get_col_index(ws, "Task Name")
    if task_name_idx:
        data_rows = [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]
        data_rows.sort(key=lambda r: str(r[task_name_idx - 1] or "").lower())
        for row_idx, row_data in enumerate(data_rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    # Step 4: Delete rows where Task Name matches list, contains "order", or "discharge summary"
    task_name_idx = get_col_index(ws, "Task Name")
    if task_name_idx:
        rows_to_delete = []
        for row in ws.iter_rows(min_row=2):
            cell_value = str(row[task_name_idx - 1].value or "").strip().lower()
            if cell_value in TASK_NAMES_TO_DELETE or "order" in cell_value or "discharge summary" in cell_value:
                rows_to_delete.append(row[0].row)
        for row_idx in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_idx)

    # Step 5a: Delete rows where Provider Last is "Williams" or "Morales"
    provider_last_idx = get_col_index(ws, "Provider Last")
    if provider_last_idx:
        rows_to_delete = []
        for row in ws.iter_rows(min_row=2):
            cell_value = str(row[provider_last_idx - 1].value or "").strip().lower()
            if cell_value in ["williams", "morales"]:
                rows_to_delete.append(row[0].row)
        for row_idx in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_idx)

    # Step 5b: Handle Status column
    status_idx = get_col_index(ws, "Status")
    if status_idx:
        rows_to_delete = []
        for row in ws.iter_rows(min_row=2):
            cell_value = str(row[status_idx - 1].value or "").strip().lower()
            if "mv" in cell_value:
                rows_to_delete.append(row[0].row)
            elif cell_value in STATUS_HIGHLIGHT:
                for cell in row:
                    cell.font = RED_FONT
                    cell.fill = RED_FILL
        for row_idx in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_idx)

    return ws

# ── PDF generation ────────────────────────────────────────
def generate_pdf(ws, save_path, date_range=""):
    pagesize = letter
    page_width, page_height = pagesize
    margin = 0.5 * inch
    usable_width = page_width - 2 * margin

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return

    headers = [str(c) if c is not None else "" for c in all_rows[0]]
    data_rows = all_rows[1:]

    status_col_idx = None
    for i, h in enumerate(headers):
        if h.strip().lower() == "status":
            status_col_idx = i
            break

    col_count = len(headers)
    task_col_idx = None
    for i, h in enumerate(headers):
        if h.strip().lower() == "task name":
            task_col_idx = i
            break

    small_col_width = 0.8 * inch
    if task_col_idx is not None:
        remaining = usable_width - (small_col_width * (col_count - 1))
        col_widths = [small_col_width] * col_count
        col_widths[task_col_idx] = remaining
    else:
        col_widths = [usable_width / col_count] * col_count

    normal_style = ParagraphStyle("cell", fontName="Helvetica", fontSize=8, leading=11, wordWrap="LTR")
    red_style = ParagraphStyle("red_cell", fontName="Helvetica", fontSize=8, leading=11, wordWrap="LTR", textColor=colors.red)
    header_style = ParagraphStyle("header", fontName="Helvetica-Bold", fontSize=8, leading=11, textColor=colors.white, wordWrap="LTR")
    title_style = ParagraphStyle("title", fontName="Helvetica-Bold", fontSize=14, leading=20, alignment=TA_CENTER)
    summary_label_style = ParagraphStyle("sum_label", fontName="Helvetica-Bold", fontSize=10, leading=14, alignment=TA_LEFT)
    summary_value_style = ParagraphStyle("sum_value", fontName="Helvetica", fontSize=10, leading=14, alignment=TA_LEFT)
    summary_red_style = ParagraphStyle("sum_red", fontName="Helvetica-Bold", fontSize=10, leading=14, alignment=TA_LEFT, textColor=colors.red)
    summary_billable_style = ParagraphStyle("sum_billable", fontName="Helvetica-Bold", fontSize=10, leading=14, alignment=TA_LEFT, textColor=colors.HexColor("#1a5276"))

    total_visits = len(data_rows)
    red_visits = 0
    if status_col_idx is not None:
        for row in data_rows:
            val = str(row[status_col_idx] or "").strip().lower()
            if val in STATUS_HIGHLIGHT:
                red_visits += 1
    billable_visits = total_visits - red_visits

    table_data = [[Paragraph(h, header_style) for h in headers]]
    for row in data_rows:
        is_red = False
        if status_col_idx is not None:
            val = str(row[status_col_idx] or "").strip().lower()
            if val in STATUS_HIGHLIGHT:
                is_red = True
        style = red_style if is_red else normal_style
        cells = [Paragraph(str(c) if c is not None else "", style) for c in row]
        table_data.append(cells)

    table_style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a5276")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cccccc")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#eaf0fb")]),
    ]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle(table_style_cmds))

    summary_data = [
        [Paragraph("Total Visits:", summary_label_style), Paragraph(str(total_visits), summary_value_style)],
        [Paragraph("Red Visits (Not Started / Saved):", summary_red_style), Paragraph(str(red_visits), summary_red_style)],
        [Paragraph("Billable Visits:", summary_billable_style), Paragraph(str(billable_visits), summary_billable_style)],
    ]
    summary_table = Table(summary_data, colWidths=[3*inch, 1*inch])
    summary_table.setStyle(TableStyle([
        ("LINEABOVE", (0, 0), (-1, 0), 1, colors.HexColor("#cccccc")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))

    elements = []
    if os.path.exists(LOGO_PATH):
        logo = RLImage(LOGO_PATH, width=0.8*inch, height=0.8*inch)
        elements.append(logo)
    elements.append(Spacer(1, 0.05*inch))
    title_text = "Schedule Report for Payroll"
    if date_range:
        title_text += f"<br/><font size=11>Schedule from {date_range}</font>"
    elements.append(Paragraph(title_text, title_style))
    elements.append(Spacer(1, 0.15*inch))
    elements.append(table)
    elements.append(Spacer(1, 0.2*inch))
    elements.append(summary_table)

    doc = SimpleDocTemplate(save_path, pagesize=pagesize, leftMargin=margin, rightMargin=margin, topMargin=margin, bottomMargin=margin)
    doc.build(elements)

# ── Upload & process ──────────────────────────────────────
def process_and_export(file_path):
    status_label.configure(text="Loading file...")
    app.update()

    workbook = openpyxl.load_workbook(file_path)
    ws = workbook.active

    row_count = ws.max_row - 1
    col_count = ws.max_column
    status_label.configure(text=f"Loaded: {row_count} rows, {col_count} columns. Processing...")
    app.update()

    status_label.configure(text="Applying rules...")
    app.update()
    ws = process_data(ws)

    dates = []
    target_date_idx = get_col_index(ws, "Target Date")
    if target_date_idx:
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = str(row[target_date_idx - 1] or "").strip()
            try:
                dates.append(datetime.strptime(val, "%m/%d/%Y"))
            except:
                pass
    if dates:
        date_range = f"{min(dates).strftime('%m/%d/%Y')} to {max(dates).strftime('%m/%d/%Y')}"
        filename = f"Schedule Report for {min(dates).strftime('%m/%d/%Y')} - {max(dates).strftime('%m/%d/%Y')}"
    else:
        date_range = ""
        filename = "schedule_report"

    clean_pdf_path = filedialog.asksaveasfilename(
        title="Save PDF as",
        defaultextension=".pdf",
        filetypes=[("PDF files", "*.pdf")],
        initialfile=f"{filename}.pdf"
    )
    if clean_pdf_path:
        status_label.configure(text="Generating PDF...")
        app.update()
        generate_pdf(ws, clean_pdf_path, date_range)

    excel_path = filedialog.asksaveasfilename(
        title="Save Excel file as",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialfile=f"{filename}.xlsx"
    )
    if excel_path:
        status_label.configure(text="Saving Excel file...")
        app.update()
        workbook.save(excel_path)

    status_label.configure(text=f"All done! Rows in final report: {ws.max_row - 1}")

def upload_file():
    file_path = filedialog.askopenfilename(
        title="Select your schedule file",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
    )
    if not file_path:
        return
    process_and_export(file_path)

def on_drag_enter(e):
    drop_zone.configure(fg_color=("gray80", "gray30"))

def on_drag_leave(e):
    drop_zone.configure(fg_color=("gray90", "gray20"))

def on_drop(e):
    drop_zone.configure(fg_color=("gray90", "gray20"))
    file_path = e.data.strip().strip('{}')
    if file_path.endswith(('.xlsx', '.xls')):
        process_and_export(file_path)
    else:
        status_label.configure(text="Please drop an Excel file (.xlsx or .xls)")

# ── Build UI ──────────────────────────────────────────────
if DND_AVAILABLE:
    app = TkinterDnD.Tk()
    app.title("Schedule Report For Payroll")
    app.geometry("800x600")
    ctk.set_appearance_mode("System")
    ctk.set_default_color_theme("blue")
else:
    app = ctk.CTk()
    app.title("Schedule Report For Payroll")
    app.geometry("800x600")

if os.path.exists(LOGO_PATH):
    pil_image = PILImage.open(LOGO_PATH)
    logo_image = ctk.CTkImage(light_image=pil_image, dark_image=pil_image, size=(120, 120))
    logo_label = ctk.CTkLabel(app, image=logo_image, text="")
    logo_label.pack(pady=(20, 5))

title_label = ctk.CTkLabel(app, text="Schedule Report For Payroll", font=ctk.CTkFont(size=24, weight="bold"))
title_label.pack(pady=(5, 0))

subtitle_label = ctk.CTkLabel(app, text="Upload a schedule, apply rules, and export a PDF and Excel file.", font=ctk.CTkFont(size=14))
subtitle_label.pack(pady=5)

version_label = ctk.CTkLabel(app, text=f"Version {CURRENT_VERSION}", font=ctk.CTkFont(size=11), text_color="gray")
version_label.pack(pady=0)

upload_btn = ctk.CTkButton(app, text="Upload Schedule", font=ctk.CTkFont(size=16), height=50, width=250, command=upload_file)
upload_btn.pack(pady=20)

drop_zone = ctk.CTkLabel(
    app,
    text="or drag and drop your Excel file here",
    font=ctk.CTkFont(size=15),
    text_color="gray",
    fg_color=("gray90", "gray20"),
    corner_radius=10,
    width=500,
    height=120
)
drop_zone.pack(pady=(0, 10))

if DND_AVAILABLE:
    drop_zone.drop_target_register(DND_FILES)
    drop_zone.dnd_bind('<<DropEnter>>', on_drag_enter)
    drop_zone.dnd_bind('<<DropLeave>>', on_drag_leave)
    drop_zone.dnd_bind('<<Drop>>', on_drop)

status_label = ctk.CTkLabel(app, text="No file loaded yet.", font=ctk.CTkFont(size=13), text_color="gray")
status_label.pack(pady=5)

threading.Thread(target=check_for_updates, daemon=True).start()

app.mainloop()